#!/usr/bin/env python3
"""
Script pour extraire les images d'un fichier Excel et les associer aux produits
"""
import openpyxl
from openpyxl.drawing.image import Image
import os
import sys
from pathlib import Path

# Configuration
EXCEL_FILE = '/Users/thierrycyrillefrancillette/Downloads/fichier en csv/ARTICLES BAZAR.xlsx'
OUTPUT_DIR = '/Users/thierrycyrillefrancillette/Downloads/images_produits'

def extract_images():
    """Extrait toutes les images du fichier Excel"""

    print("🚀 Extraction des images du fichier Excel...")
    print(f"📁 Fichier: {EXCEL_FILE}")
    print()

    # Créer le dossier de sortie
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Charger le workbook
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"❌ Erreur lors de l'ouverture du fichier: {e}")
        return

    ws = wb.active
    print(f"📋 Feuille active: {ws.title}")
    print(f"📊 Nombre de lignes: {ws.max_row}")
    print()

    # Lire les SKU de la colonne "REF ARTICLE"
    skus = {}
    for row in range(2, ws.max_row + 1):  # Commencer à la ligne 2 (après l'en-tête)
        cell_value = ws.cell(row=row, column=2).value  # Colonne B = REF ARTICLE
        if cell_value:
            skus[row] = str(cell_value).strip()

    print(f"✅ {len(skus)} SKU trouvés")
    print()

    # Extraire les images
    if not hasattr(ws, '_images') or not ws._images:
        print("⚠️  Aucune image trouvée dans le fichier Excel")
        print()
        print("💡 Astuce: Les images doivent être intégrées dans les cellules,")
        print("   pas seulement copiées-collées.")
        return

    images_count = 0
    print(f"🖼️  {len(ws._images)} images trouvées dans la feuille")
    print()

    for idx, img in enumerate(ws._images):
        try:
            # Récupérer l'ancre de l'image (position)
            anchor = img.anchor

            # Déterminer la ligne de l'image
            if hasattr(anchor, '_from'):
                row = anchor._from.row + 1  # +1 car openpyxl commence à 0
            elif hasattr(anchor, 'row'):
                row = anchor.row + 1
            else:
                # Si on ne peut pas déterminer la ligne, utiliser un nom générique
                filename = f"image_{idx + 1}.png"
                filepath = os.path.join(OUTPUT_DIR, filename)
                # Sauvegarder l'image à partir des données binaires
                with open(filepath, 'wb') as f:
                    f.write(img.ref.getvalue())
                images_count += 1
                print(f"  ✓ Image {idx + 1} sauvegardée: {filename}")
                continue

            # Récupérer le SKU correspondant
            sku = skus.get(row, f"unknown_{idx}")

            # Nettoyer le SKU pour le nom de fichier
            safe_sku = "".join(c for c in sku if c.isalnum() or c in ('-', '_')).strip()

            # Sauvegarder l'image
            filename = f"{safe_sku}.png"
            filepath = os.path.join(OUTPUT_DIR, filename)

            # Si le fichier existe déjà, ajouter un numéro
            counter = 1
            while os.path.exists(filepath):
                filename = f"{safe_sku}_{counter}.png"
                filepath = os.path.join(OUTPUT_DIR, filename)
                counter += 1

            # Sauvegarder l'image à partir des données binaires
            with open(filepath, 'wb') as f:
                f.write(img.ref.getvalue())
            images_count += 1
            print(f"  ✓ Image {images_count}: {filename} (ligne {row}, SKU: {sku})")

        except Exception as e:
            print(f"  ❌ Erreur lors de l'extraction de l'image {idx + 1}: {e}")

    print()
    print("=" * 70)
    print(f"✅ Extraction terminée!")
    print(f"📦 {images_count} images extraites")
    print(f"📁 Dossier: {OUTPUT_DIR}")
    print("=" * 70)

if __name__ == "__main__":
    extract_images()
